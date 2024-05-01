import openpyxl as xl
from openpyxl.styles import Font

# create new excel doc
wb = xl.Workbook()

ws = wb.active

ws.title = 'First Sheet'

wb.create_sheet(index=1,title="Second Sheet")

# write content to a cell

ws['A1'] = 'Invoice'

headerfont = Font(name = 'Times New Roman', size = 24, bold = True)

ws['A1'].font = headerfont

ws['A2'] = 'Tires'
ws['A3'] = 'Brakes'
ws['A4'] = 'Alignment'

ws.merge_cells('A1:B1')
# To unmerge cells
# ws.unmerge_cells('A1:B1')

ws['B2'] = 450
ws['B3'] = 225
ws['B4'] = 150

ws['A8'] = 'Total'
ws['A8'].font = Font(size=16,bold=True)

ws.column_dimensions['A'].width = 25

ws['B8'] = '=SUM(B2:B4)'

# Read the excel file - 'ProduceReport.xlsx'
# Write all the content of this file to 'Second Sheet'
# Display Grand Total and Avg of 'Amt Sold' and 'Total'

write_sheet = wb['Second Sheet']

read_wb = xl.load_workbook('ProduceReport.xlsx')
read_ws = read_wb['ProduceReport']

# for current_row in read_ws.iter_rows(min_row=1, max_row=read_ws.max_row, max_col=read_ws.max_column):

# Class example
for row in read_ws.iter_rows():
    ls = [i.value for i in row]
    write_sheet.append(ls)

max_row = write_sheet.max_row

write_sheet.cell(max_row+2,2).value = 'Total'
write_sheet.cell(max_row+2,2).font = Font(size=16,bold=True)

write_sheet.cell(max_row+4,2).value = 'Average'
write_sheet.cell(max_row+4,2).font = Font(size=16,bold=True)

write_sheet.cell(max_row+2,3).value = '=SUM(C2:C' + str(max_row) +')'
write_sheet.cell(max_row+2,4).value = '=SUM(D2:D' + str(max_row) +')'

write_sheet.cell(max_row+4,3).value = '=AVERAGE(C2:C' + str(max_row) +')'
write_sheet.cell(max_row+4,4).value = '=AVERAGE(D2:D' + str(max_row) +')'

write_sheet.column_dimensions['A'].width = 16
write_sheet.column_dimensions['B'].width = 15
write_sheet.column_dimensions['C'].width = 15
write_sheet.column_dimensions['D'].width = 15

for cell in write_sheet["C:C"]:
    cell.number_format = '#,##0'

for cell in write_sheet['D:D']:
    cell.number_format = u'"$ "#,##0.00'

# Attempt 1
'''
for current_row in read_ws['A1':'D41']:
    for current_cell in current_row:
        write_sheet[current_cell.coordinate] = current_cell.value

write_sheet['B43'] = 'Total'

write_sheet['B45'] = 'Average'

write_sheet['C43'] = '=SUM(C2:C41)'

write_sheet['D43'] = '=SUM(D2:D41)'

write_sheet['C45'] = write_sheet['C43'].value / 

write_sheet['D45'] = 
'''

'''
amt_sold_total = 0
for row in range(2, write_sheet.max_row+1):
    amt_sold_total += (write_sheet.cell(row,3).value)
grand_total = 0
for row in range(2, write_sheet.max_row+1):

'''


wb.save("PythontoExcel.xlsx")


