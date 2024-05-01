
from urllib.request import urlopen
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font





#webpage = 'https://www.boxofficemojo.com/weekend/chart/'
webpage = 'https://www.boxofficemojo.com/year/2024/'

page = urlopen(webpage)			

soup = BeautifulSoup(page, 'html.parser')

title = soup.title

print(title.text)

table_rows = soup.findAll('tr')

wb = xl.Workbook()

ws = wb.active

ws.title = 'Box Office Report'

headers = ["No.", "Movie Title", "Release Date", "Number of Theaters", "Total Gross", "Average Gross by Theater"]

ws.append(headers)

row = 1

for entry in table_rows[1:6]:
    td = entry.findAll("td")
    row += 1
    number = td[0].text
    movie = td[1].text
    date = td[8].text
    number_of_theaters = int(td[6].text.replace(",",""))
    total_gross = int(td[7].text.strip('$').replace(",",""))

    '''
    testing values
    
    print(number)
    print(movie)
    print(number_of_theaters)
    print(date)
    print(total_gross)
    print()
    print(row)
    '''

    ws.cell(row,1).value = number
    ws.cell(row,2).value = movie
    ws.cell(row,3).value = date
    ws.cell(row,4).value = number_of_theaters
    ws.cell(row,5).value = total_gross
    ws.cell(row,6).value = '=E'+str(row)+'/D'+str(row)

wb.save("BoxOfficeReport.xlsx")

##
##
##
##

